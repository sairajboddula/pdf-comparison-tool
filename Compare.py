from openpyxl import Workbook
import os.path, os, filecmp, openpyxl, re
from pypdf import PdfReader
from datetime import datetime
import hashlib, json


now = datetime.now()
dt = now.strftime("%y-%m-%d-%H-%M-%S")
wb= Workbook()
ws = wb.active
ws['A1']="Dir_Org"
ws['B1']="Dir_Comp"
ws['C1']="Matched"
ws['D1']="Error Place"
ws['E1']="Original"
ws['F1']="Changed"
ws['G1']="Checksum Base"
ws['H1']="CheckSum Generated"

ws.title="pdf"

# Add a new sheet for folder summary
ws_folder = wb.create_sheet("FolderSummary")
ws_folder['A1'] = "Subfolder Name"
ws_folder['B1'] = "Baseline"
ws_folder['C1'] = "Generated"
ws_folder['D1'] = "Difference"


ws_file = wb.create_sheet("File Error")
ws_file['A1'] = "File"
ws_file['B1'] = "Error"
with open("ignore.json", "r") as f:
    config = json.load(f)
y_min = config['pdf_py']['coordinates']['y_min']
y_max = config['pdf_py']['coordinates']['y_max']

def Compare(dir_path,dir_path1,output):
    if not os.path.exists(dir_path) or not os.path.exists(dir_path1):
        raise ValueError(f"Invalid input directories: {dir_path} or {dir_path1} does not exist.")
    c = 0
    subfolder_names_dir1 = set(os.path.basename(root) for root, dirs, files in os.walk(dir_path))
    subfolder_names_dir2 = set(os.path.basename(root) for root, dirs, files in os.walk(dir_path1))

    common_subfolders = list(set(subfolder_names_dir1) & set(subfolder_names_dir2))
    filelist1=[]
    filelist2=[]
    filelist3=[]
    filelist4=[]

    empty_pdf_files = []
    common_files = []
    extra_files = []

    r = 2
    for subfolder in common_subfolders:
        subfolder_path1 = os.path.join(dir_path, subfolder)
        subfolder_path2 = os.path.join(dir_path1, subfolder)
        files1 = [os.path.join(root, file) for root, dirs, files in os.walk(subfolder_path1) for file in files if file.endswith('.pdf')]
        files2 = [os.path.join(root, file) for root, dirs, files in os.walk(subfolder_path2) for file in files if file.endswith('.pdf')]
        # File paths of pdf files in first directory
        file_count = len(files1)
        file_count1 = len(files2)
        ws_folder.cell(row=r, column=1).value = subfolder
        ws_folder.cell(row=r, column=2).value = file_count
        ws_folder.cell(row=r, column=3).value = file_count1
        
        

        for root1, dirs, files in os.walk(subfolder_path1):
            for file in files:
                #append the file name to the list
                filelist1.append(os.path.join(root1,file))
        for root2, dirs, files in os.walk(subfolder_path2):
            for file in files:
                #append the file name to the list
                filelist2.append(os.path.join(root2,file))
        for root3, dirs, files in os.walk(subfolder_path1):
            for file in files:
                #append the file name to the list
                filelist3.append(file)
        for root4, dirs, files in os.walk(subfolder_path2):
            for file in files:
                #append the file name to the list
                filelist4.append(file)      
        
        file1=[]
        file2=[]
        file3=[]
        file4=[]
        for fil in filelist1:
            if fil.endswith('.pdf'):
                file1.append(fil)
    
        for fil1 in filelist2:
            if fil1.endswith('.pdf'):
                file2.append(fil1)

        for fil2 in filelist3:
            if fil2.endswith('.pdf'):
                file3.append(fil2)
                if os.path.exists(os.path.join(subfolder_path1, fil2)):
                    if os.path.getsize(os.path.join(subfolder_path1, fil2))==0:
                        empty_pdf_files.append(fil2)

        for fil3 in filelist4:
            if fil3.endswith('.pdf'):
                file4.append(fil3)
                if os.path.exists(os.path.join(subfolder_path2, fil3)):
                    if os.path.getsize(os.path.join(subfolder_path2, fil3))==0:
                        empty_pdf_files.append(fil3)


        if file_count == file_count1:
            diffe = "None"
            ws_folder.cell(row=r, column=4).value = diffe
        else:
            diffe = list(set(file3).symmetric_difference(set(file4)))
            ws_folder.cell(row=r, column=4).value = ",".join(diffe)
        r+=1


    if len(file1) == len(file2):
        # return True
        p1=2
        for j in range(len(file1)):
            c+=1
            if os.path.basename(os.path.dirname(file1[j])) == os.path.basename(os.path.dirname(file2[j])) and file3[j] == file4[j]:
                reader = PdfReader(file1[j])
                reader1 = PdfReader(file2[j])
                
                num = len(reader.pages)
                num1 = len(reader1.pages)
                if (num == num1)==0 or (num!=num1):
                    ws_file.cell(row=p1, column=1).value = file1[j]
                    ws_file.cell(row=p1, column=2).value = "Pages are not same/empty"
                    p1 += 1
                    continue
                else:
                    f1 = open(output + os.sep +'\\filename.txt', 'w',encoding='utf8')
                    f2 = open(output + os.sep +'\\filename1.txt', 'w',encoding='utf8')
                    for i in range (0,num):
                        page = reader.pages[i]

                        parts = []

                        def visitor_body(text, cm, tm, fontDict, fontSize):
                            y = tm[5]
                            if y > y_min and y < y_max:
                                parts.append(text)


                        page.extract_text(visitor_text=visitor_body)
                        text_body = "".join(parts)

                        print(text_body,file=f1)
                    for i in range (0,num1):
                        page = reader1.pages[i]

                        parts = []

                        def visitor_body(text, cm, tm, fontDict, fontSize):
                            y = tm[5]
                            if y >y_min and y < y_max:
                                parts.append(text)


                        page.extract_text(visitor_text=visitor_body)
                        text_body = "".join(parts)

                        print(text_body,file=f2)

                    f1.close()                                       
                    f2.close()
                    temp1 = output + os.sep +'\\filename.txt'
                    temp2 = output + os.sep +'\\filename1.txt'
                    f1 = open(temp1, "r",encoding='utf8')  
                    f2 = open(temp2, "r",encoding='utf8')  
                    d= output + os.sep +"\\output"+str(c)+".txt"
                    f3 = open(d, "w",encoding='utf8')
                    i = 0

                    result = filecmp.cmp(temp1,temp2)
                    if result == True:
                        print("Both files are Equal",file=f3)
                        e = 0
                        f1.close()                                       
                        f2.close()
                        os.remove(temp1)
                        os.remove(temp2)
                    
                    else:
                        e= 1
                        for line1 in f1:
                            i += 1
                            
                            for line2 in f2:
                                
                                # matching line1 from both files
                                if line1 == line2:  
                                    # print IDENTICAL if similar
                                    # print("Line ", i, ": IDENTICAL",file=f3)   
                                    pass    
                                else:
                                    for i3 in range(0,num):
                                        pbob = reader.pages[i3]
                                        text3 = pbob.extract_text()
                                        if re.search(line1,text3):
                                            z = i3 + 1
                                    print("Page ",z,"| Line ", i,file=f3)
                                    # print("Line ", i, ":",file=f3)
                                    # else print that line from both files
                                    print(line1, end='',file=f3)
                                    print(line2, end='',file=f3)
                                break
                            # print(line1,file=f3)
                        # closing files
                        f1.close()                                       
                        f2.close()
                        os.remove(temp1)
                        os.remove(temp2)
                    d= output + os.sep +"\\output"+str(c)+".txt"
                    f3 = open(d, "w",encoding='utf8')
                    # rf = pd.read_csv(output + os.sep +"\\output" +str(c)+".txt")
                    # rf.to_csv(output+os.sep+"\\text"+str(c)+".csv",index=None)
                    nm=[]
                    dummy = []
                    k=0
                    with open(output+os.sep+"\\output"+str(c)+".txt", 'r',encoding="utf8") as f:
                        mn = [line.strip() for line in f]
                    # print(mn)
                    for i in mn:
                        k+=1
                        dummy.append(i)
                        if k == 3:
                            nm.append(dummy)
                            k=0
                            dummy=[]
                    dummy=[('',)]
                    # mn.append(nm)
                    a = str(file1[j])
                    b = str(file2[j])
                    # f = str(exceldir)
                    y=[
                        (a,b,e),
                        
                    ]  
                    y = y  + dummy
                    for j in y:
                        ws.append(j)
                    column = ws['C']
                    k = len(column) -1
                    p = ws.cell(k,3).internal_value
                    with open(a, 'rb') as w:
                        pdf1 = w.read()
                        checksum = hashlib.md5(pdf1).hexdigest()
                        ws.cell(k,7).value = checksum
                    # for y2 in file2:
                    with open(b, 'rb') as w1:
                        pdf2 = w1.read()
                        checksum1 = hashlib.md5(pdf2).hexdigest()
                        ws.cell(k,8).value = checksum1
                    if p == 1:
                        for o in nm:
                            ws.cell(k, 4).value = o[0]
                            ws.cell(k, 5).value = o[1]
                            ws.cell(k, 6).value = o[2]
                            k+=1
                    
                    wb.save(filename = output+os.sep+'\\pdf-'+str(dt)+'.xlsx')
                    f3.close()
                    os.remove(d)
            else:
                continue

    else:
        # return "Files are not same in both the folders"
        

        set_file1 = set(file3)
        set_file2 = set(file4)
        p1=2
        for file in set_file1.intersection(set_file2):
            index1 = file3.index(file)
            index2 = file4.index(file)
            reader = PdfReader(file1[index1])
            reader1 = PdfReader(file2[index2])
            # compare the contents of the two files as before
            num = len(reader.pages)
            num1 = len(reader1.pages)
            if (num == num1)==0 or (num!=num1):
                    ws_file.cell(row=p1, column=1).value = file1[index1]
                    ws_file.cell(row=p1, column=2).value = "Pages are not same/empty"
                    p1 += 1
                    continue
            else:
                f1 = open(output + os.sep +'\\filename.txt', 'w',encoding='utf8')
                f2 = open(output + os.sep +'\\filename1.txt', 'w',encoding='utf8')
                for i in range (0,num):
                    page = reader.pages[i]

                    parts = []

                    def visitor_body(text, cm, tm, fontDict, fontSize):
                        y = tm[5]
                        if y > y_min and y < y_max:
                            parts.append(text)


                    page.extract_text(visitor_text=visitor_body)
                    text_body = "".join(parts)

                    print(text_body,file=f1)
                for i in range (0,num1):
                    page = reader1.pages[i]

                    parts = []

                    def visitor_body(text, cm, tm, fontDict, fontSize):
                        y = tm[5]
                        if y > y_min and y < y_max:
                            parts.append(text)


                    page.extract_text(visitor_text=visitor_body)
                    text_body = "".join(parts)

                    print(text_body,file=f2)

                f1.close()                                       
                f2.close()
                temp1 = output + os.sep +'\\filename.txt'
                temp2 = output + os.sep +'\\filename1.txt'
                f1 = open(temp1, "r",encoding='utf8')  
                f2 = open(temp2, "r",encoding='utf8')  
                d= output + os.sep +"\\output"+str(c)+".txt"
                f3 = open(d, "w",encoding='utf8')
                i = 0

                result = filecmp.cmp(temp1,temp2)
                if result == True:
                    print("Both files are Equal",file=f3)
                    e = 0
                    f1.close()                                       
                    f2.close()
                    os.remove(temp1)
                    os.remove(temp2)
                    
                else:
                    e=1
                    for line1 in f1:
                        i += 1
                        
                        for line2 in f2:
                            
                            # matching line1 from both files
                            if line1 == line2:  
                                # print IDENTICAL if similar
                                # print("Line ", i, ": IDENTICAL",file=f3)   
                                pass    
                            else:
                                for i3 in range(num):
                                    pbob = reader.pages[i3]
                                    text3 = pbob.extract_text()
                                    if re.search(line1,text3):
                                        z = i3 + 1
                                print("Page ",z,"| Line ", i,file=f3)
                                # else print that line from both files
                                print(line1, end='',file=f3)
                                print(line2, end='',file=f3)
                            break
                        # print(line1,file=f3)
                    # closing files
                    f1.close()                                       
                    f2.close()
                    os.remove(temp1)
                    os.remove(temp2)
            d= output + os.sep +"\\output"+str(c)+".txt"
            f3 = open(d, "w",encoding='utf8')
            # rf = pd.read_csv(output + os.sep +"\\output" +str(c)+".txt")
            # rf.to_csv(output+os.sep+"\\text"+str(c)+".csv",index=None)
            nm=[]
            dummy = []
            k=0
            with open(output+os.sep+"\\output"+str(c)+".txt", 'r',encoding="utf8") as f:
                mn = [line.strip() for line in f]

            for i in mn:
                k+=1
                dummy.append(i)
                if k == 3:
                    nm.append(dummy)
                    k=0
                    dummy=[]
                    
            dummy=[('',)]
            
            a = str(file1[index1])
            b = str(file2[index2])
            # print("len(nm)",len(nm))

            y=[
                (a,b,e),
                
            ]
            
            y = y  + dummy
            for j in y:
                ws.append(j)
            colum = ws['C']
            k = len(colum) - 1
            # print(k)
            
            p = ws.cell(k,3).internal_value
            # for y1 in file1:
            with open(a, 'rb') as w:
                pdf1 = w.read()
                checksum = hashlib.md5(pdf1).hexdigest()
                ws.cell(k,7).value = checksum
            # for y2 in file2:
            with open(b, 'rb') as w1:
                pdf2 = w1.read()
                checksum1 = hashlib.md5(pdf2).hexdigest()
                ws.cell(k,8).value = checksum1

            if p == 1:
                for o in nm:
                    ws.cell(k, 4).value = o[0]
                    ws.cell(k, 5).value = o[1]
                    ws.cell(k, 6).value = o[2]
                    k+=1
            wb.save(filename = output+os.sep+'\\pdf-'+str(dt)+'.xlsx')
            f3.close()
            os.remove(d)
    return "Executed, Check Result folder "


a = "Enter your base line location"
b = "replace_with_generated_location"
outpt = "Output_Location"
comp = Compare(a,b,outpt)
print(comp)